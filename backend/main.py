"""FastAPI application: XER import, CPM, diagnostics, exports."""

from __future__ import annotations

import csv
import sqlite3
from pathlib import Path

from dotenv import load_dotenv

# Load backend/.env regardless of the process working directory (e.g. uvicorn from repo root).
load_dotenv(Path(__file__).resolve().parent / ".env")
import io
from datetime import datetime, timedelta
from typing import Any, List, Optional

from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ai_routes import router as ai_router
from cpm_engine import run_cpm_for_project_rows
from database import DEFAULT_DB_PATH, get_connection, init_db
from deps import fetch_activities, fetch_relationships, get_db
from diagnostics import run_diagnostics
from models import (
    ActivitiesPage,
    ActivityRow,
    BaselineComparison,
    BaselineInfo,
    CpmResult,
    DiagnosticsResult,
    ImportResult,
    ProjectSummary,
    RelationshipRow,
    WbsCreateBody,
    WbsUpdateBody,
)
from calendar_engine import load_calendars, hours_to_calendar_date, get_default_calendar
from xer_parser import import_xer_to_sqlite

app = FastAPI(title="Primavera P6 XER Local Analyzer", version="1.0.0")

app.include_router(ai_router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://127.0.0.1:3000",
        "http://localhost:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup() -> None:
    """Ensure schema exists on process start."""
    conn = get_connection(DEFAULT_DB_PATH)
    init_db(conn)
    conn.close()


@app.post("/api/import", response_model=ImportResult)
async def import_xer(file: UploadFile = File(...)) -> ImportResult:
    if not file.filename or not file.filename.lower().endswith(".xer"):
        raise HTTPException(400, "Upload a .xer file")
    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty file")
    conn = get_db()
    try:
        summary = import_xer_to_sqlite(conn, data, filename=file.filename)
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    finally:
        conn.close()
    return ImportResult(
        proj_id=summary["proj_id"],
        name=summary["name"],
        activities_count=summary["activities_count"],
        relationships_count=summary["relationships_count"],
        calendars_count=summary.get("calendars_count", 0),
        wbs_count=summary.get("wbs_count", 0),
        resources_count=summary.get("resources_count", 0),
        task_resources_count=summary.get("task_resources_count", 0),
    )


@app.get("/api/projects", response_model=List[ProjectSummary])
def list_projects() -> List[ProjectSummary]:
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            """SELECT p.proj_id, p.name, p.imported_at,
               (SELECT COUNT(*) FROM activities a WHERE a.proj_id = p.proj_id),
               (SELECT COUNT(*) FROM relationships r WHERE r.proj_id = p.proj_id)
               FROM projects p ORDER BY p.imported_at DESC"""
        )
        rows = cur.fetchall()
        return [
            ProjectSummary(
                proj_id=r[0],
                name=r[1],
                imported_at=r[2],
                activity_count=r[3],
                relationship_count=r[4],
            )
            for r in rows
        ]
    finally:
        conn.close()


def _sort_activities(
    items: List[ActivityRow],
    sort_by: Optional[str],
    sort_dir: str,
) -> List[ActivityRow]:
    if not sort_by:
        return items
    rev = sort_dir.lower() == "desc"
    key_map = {
        "task_id": lambda x: x.task_id,
        "name": lambda x: (x.name or "").lower(),
        "duration_hrs": lambda x: x.duration_hrs,
        "early_start": lambda x: x.early_start if x.early_start is not None else -1e30,
        "early_finish": lambda x: x.early_finish if x.early_finish is not None else -1e30,
        "late_start": lambda x: x.late_start if x.late_start is not None else -1e30,
        "late_finish": lambda x: x.late_finish if x.late_finish is not None else -1e30,
        "total_float_hrs": lambda x: x.total_float_hrs if x.total_float_hrs is not None else -1e30,
    }
    fn = key_map.get(sort_by, key_map["task_id"])
    return sorted(items, key=fn, reverse=rev)


@app.get("/api/projects/{proj_id}/activities", response_model=ActivitiesPage)
def get_activities(
    proj_id: str,
    search: Optional[str] = None,
    critical_only: bool = False,
    sort_by: Optional[str] = None,
    sort_dir: str = "asc",
) -> ActivitiesPage:
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        rows = fetch_activities(conn, proj_id)
    finally:
        conn.close()

    if sort_dir not in ("asc", "desc"):
        sort_dir = "asc"

    items: List[ActivityRow] = []
    for r in rows:
        if critical_only and not int(r.get("is_critical") or 0):
            continue
        if search:
            s = search.lower()
            if (
                s not in (r.get("name") or "").lower()
                and s not in str(r.get("task_id"))
                and s not in str(r.get("task_code") or "").lower()
            ):
                continue
        items.append(
            ActivityRow(
                proj_id=r["proj_id"],
                task_id=str(r["task_id"]),
                task_code=r.get("task_code") or str(r["task_id"]),
                name=r.get("name") or "",
                duration_hrs=float(r.get("duration_hrs") or 0),
                wbs_id=r.get("wbs_id"),
                wbs_name=r.get("wbs_name"),
                wbs_short_name=r.get("wbs_short_name"),
                calendar_id=r.get("calendar_id"),
                early_start=r.get("early_start"),
                early_finish=r.get("early_finish"),
                late_start=r.get("late_start"),
                late_finish=r.get("late_finish"),
                total_float_hrs=r.get("total_float_hrs"),
                free_float_hrs=r.get("free_float_hrs"),
                is_critical=bool(int(r.get("is_critical") or 0)),
                is_near_critical=bool(int(r.get("is_near_critical") or 0)),
                is_milestone=bool(int(r.get("is_milestone") or 0)),
                constraint_type=r.get("constraint_type"),
                constraint_date=r.get("constraint_date"),
                actual_start=r.get("actual_start"),
                actual_finish=r.get("actual_finish"),
                remaining_duration_hrs=r.get("remaining_duration_hrs"),
                percent_complete=r.get("percent_complete") or 0,
            )
        )
    items = _sort_activities(items, sort_by, sort_dir)
    return ActivitiesPage(items=items, total=len(items))


@app.get("/api/projects/{proj_id}/relationships", response_model=List[RelationshipRow])
def get_relationships(proj_id: str) -> List[RelationshipRow]:
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        raw = fetch_relationships(conn, proj_id)
    finally:
        conn.close()
    return [
        RelationshipRow(
            id=int(r["id"]),
            pred_id=str(r["pred_id"]),
            succ_id=str(r["succ_id"]),
            rel_type=str(r["rel_type"]),
            lag_hrs=float(r.get("lag_hrs") or 0),
        )
        for r in raw
    ]


@app.get("/api/projects/{proj_id}/wbs")
def get_wbs(proj_id: str) -> List[dict]:
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        cur.execute(
            """SELECT wbs_id, parent_wbs_id, wbs_short_name, wbs_name, seq_num
               FROM wbs WHERE proj_id = ? ORDER BY seq_num, wbs_id""",
            (proj_id,),
        )
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]
    finally:
        conn.close()


@app.post("/api/projects/{proj_id}/wbs")
def create_wbs(proj_id: str, body: WbsCreateBody) -> dict:
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        wbs_id = str(body.wbs_id).strip()
        if not wbs_id:
            raise HTTPException(400, "wbs_id required")
        wbs_name = body.wbs_name if body.wbs_name is not None else f"WBS {wbs_id}"
        try:
            cur.execute(
                """INSERT INTO wbs (wbs_id, proj_id, parent_wbs_id, wbs_short_name, wbs_name, seq_num)
                   VALUES (?, ?, ?, ?, ?, 0)""",
                (wbs_id, proj_id, body.parent_wbs_id, body.wbs_short_name or "", wbs_name),
            )
            conn.commit()
        except sqlite3.IntegrityError as e:
            if "UNIQUE" in str(e).upper():
                raise HTTPException(409, "WBS already exists") from e
            raise HTTPException(400, str(e)) from e
        return {"ok": True, "wbs_id": wbs_id}
    finally:
        conn.close()


@app.put("/api/projects/{proj_id}/wbs/{wbs_id}")
def update_wbs(proj_id: str, wbs_id: str, body: WbsUpdateBody) -> dict:
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        sets: List[str] = []
        vals: List[Any] = []
        data = body.model_dump(exclude_unset=True)
        for k in ("wbs_name", "wbs_short_name", "parent_wbs_id"):
            if k in data:
                sets.append(f"{k} = ?")
                vals.append(data[k])
        if not sets:
            raise HTTPException(400, "No fields to update")
        vals.extend([proj_id, wbs_id])
        cur.execute(f"UPDATE wbs SET {', '.join(sets)} WHERE proj_id = ? AND wbs_id = ?", vals)
        if cur.rowcount == 0:
            raise HTTPException(404, "WBS not found")
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@app.delete("/api/projects/{proj_id}")
def delete_project(proj_id: str) -> dict:
    """Delete a project and all related data."""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        cur.execute("DELETE FROM baseline_activities WHERE baseline_id IN (SELECT id FROM baselines WHERE proj_id = ?)", (proj_id,))
        cur.execute("DELETE FROM baselines WHERE proj_id = ?", (proj_id,))
        cur.execute("DELETE FROM xer_raw_tables WHERE proj_id = ?", (proj_id,))
        cur.execute("DELETE FROM task_activity_codes WHERE proj_id = ?", (proj_id,))
        cur.execute("DELETE FROM task_resources WHERE proj_id = ?", (proj_id,))
        cur.execute("DELETE FROM resources WHERE proj_id = ?", (proj_id,))
        cur.execute("DELETE FROM relationships WHERE proj_id = ?", (proj_id,))
        cur.execute("DELETE FROM activities WHERE proj_id = ?", (proj_id,))
        cur.execute("DELETE FROM calendars WHERE proj_id = ?", (proj_id,))
        cur.execute("DELETE FROM wbs WHERE proj_id = ?", (proj_id,))
        cur.execute("DELETE FROM projects WHERE proj_id = ?", (proj_id,))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@app.post("/api/projects/{proj_id}/cpm", response_model=CpmResult)
def run_cpm(proj_id: str) -> CpmResult:
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        acts = fetch_activities(conn, proj_id)
        rels = fetch_relationships(conn, proj_id)
        err, results, proj_end, critical_path = run_cpm_for_project_rows(acts, rels)
        if err:
            return CpmResult(
                proj_id=proj_id,
                project_end_hrs=0.0,
                critical_count=0,
                total_count=0,
                critical_path=[],
                cycle_error=err,
            )
        crit_ct = sum(1 for v in results.values() if v.get("is_critical"))
        for tid, vals in results.items():
            conn.execute(
                """UPDATE activities SET
                   early_start = ?, early_finish = ?, late_start = ?, late_finish = ?,
                   total_float_hrs = ?, free_float_hrs = ?, is_critical = ?, is_near_critical = ?
                   WHERE proj_id = ? AND task_id = ?""",
                (
                    vals["early_start"],
                    vals["early_finish"],
                    vals["late_start"],
                    vals["late_finish"],
                    vals["total_float_hrs"],
                    vals["free_float_hrs"],
                    vals["is_critical"],
                    vals["is_near_critical"],
                    proj_id,
                    tid,
                ),
            )
        conn.commit()
        return CpmResult(
            proj_id=proj_id,
            project_end_hrs=proj_end,
            critical_count=crit_ct,
            total_count=len(results),
            critical_path=critical_path,
            cycle_error=None,
        )
    finally:
        conn.close()


@app.get("/api/projects/{proj_id}/diagnostics", response_model=DiagnosticsResult)
def get_diagnostics(
    proj_id: str,
    hours_per_day: float = Query(8.0, ge=0.1, le=24.0),
) -> DiagnosticsResult:
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        acts = fetch_activities(conn, proj_id)
        rels = fetch_relationships(conn, proj_id)
    finally:
        conn.close()
    return run_diagnostics(proj_id, acts, rels, hours_per_day=hours_per_day)


def _hours_to_date_str(hrs: Optional[float], ref_ms: int = None) -> str:
    """Convert hours from ref point to DD-MMM-YY format."""
    from constants import REF_MS as _REF_MS
    if hrs is None:
        return ""
    if ref_ms is None:
        ref_ms = _REF_MS
    from datetime import datetime as _dt, timezone as _tz
    ms = ref_ms + int(hrs * 3600000)
    dt = _dt.fromtimestamp(ms / 1000, tz=_tz.utc)
    return dt.strftime("%d-%b-%y").upper()


def _hours_to_days(hrs: Optional[float], hours_per_day: float = None) -> Optional[float]:
    """Convert hours to working days."""
    from constants import HOURS_PER_DAY as _HPD
    if hrs is None:
        return None
    if hours_per_day is None:
        hours_per_day = _HPD
    return round(hrs / hours_per_day, 2)


@app.get("/api/projects/{proj_id}/export/activities.xlsx")
def export_activities_xlsx(proj_id: str) -> Response:
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT name FROM projects WHERE proj_id = ?", (proj_id,))
        proj_row = cur.fetchone()
        if not proj_row:
            raise HTTPException(404, "Project not found")
        proj_name = proj_row[0] or proj_id
        acts = fetch_activities(conn, proj_id)
        # Build WBS name map
        cur.execute("SELECT wbs_id, wbs_name, wbs_short_name FROM wbs WHERE proj_id = ?", (proj_id,))
        wbs_map = {}
        for r in cur.fetchall():
            wbs_map[str(r[0])] = r[1] or r[2] or str(r[0])
    finally:
        conn.close()

    from constants import HOURS_PER_DAY as HPD, REF_DATETIME
    REF = REF_DATETIME.replace(tzinfo=None)

    def h2date(h):
        if h is None:
            return ""
        return (REF + timedelta(hours=float(h))).strftime("%d-%b-%y")

    def h2days(h):
        if h is None:
            return ""
        return int(round(float(h) / HPD))

    wb = Workbook()
    ws = wb.active
    ws.title = "CPM Schedule"

    # Styles
    title_font = Font(name="Arial", bold=True, size=14, color="1F3864")
    hdr_font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    wbs_font = Font(name="Arial", bold=True, size=9)
    wbs_fill = PatternFill("solid", fgColor="FFD966")
    crit_font = Font(name="Arial", size=9, color="CC0000")
    norm_font = Font(name="Arial", size=9)
    bdr = Border(
        left=Side("thin", "CCCCCC"),
        right=Side("thin", "CCCCCC"),
        top=Side("thin", "CCCCCC"),
        bottom=Side("thin", "CCCCCC"),
    )

    # Title
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"CRITICAL PATH BASE LINE CPM SCHEDULE {proj_id} - {proj_name}"
    c.font = title_font
    c.alignment = Alignment(horizontal="center")

    # Headers row 3
    hdrs = [
        "Activity ID",
        "Activity Name",
        "Original\nDuration",
        "Remaining\nDuration",
        "Start",
        "Finish",
        "Late Start",
        "Late Finish",
        "Total\nFloat",
        "Physical %\nComplete",
    ]
    for i, h in enumerate(hdrs, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr
    ws.row_dimensions[3].height = 28

    # Column widths
    for i, w in enumerate([12, 58, 9, 9, 12, 12, 12, 12, 7, 9], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sort by WBS then task_id
    sorted_acts = sorted(acts, key=lambda a: (str(a.get("wbs_id") or ""), str(a.get("task_id") or "")))
    row_num = 4
    cur_wbs = None

    for a in sorted_acts:
        wbs = str(a.get("wbs_id") or "")
        if wbs != cur_wbs:
            cur_wbs = wbs
            wbs_name = wbs_map.get(wbs, wbs) if wbs else "(No WBS)"
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
            cell = ws.cell(row=row_num, column=1, value=wbs_name)
            cell.font = wbs_font
            cell.fill = wbs_fill
            cell.border = bdr
            row_num += 1

        is_crit = int(a.get("is_critical") or 0)
        font = crit_font if is_crit else norm_font
        rem_dur = (
            a.get("remaining_duration_hrs")
            if a.get("remaining_duration_hrs") is not None
            else a.get("duration_hrs")
        )

        vals = [
            a.get("task_code") or a.get("task_id"),
            a.get("name"),
            h2days(a.get("duration_hrs")),
            h2days(rem_dur),
            h2date(a.get("early_start")),
            h2date(a.get("early_finish")),
            h2date(a.get("late_start")),
            h2date(a.get("late_finish")),
            h2days(a.get("total_float_hrs")) if a.get("total_float_hrs") is not None else "",
            f"{float(a.get('percent_complete') or 0):.0f}%",
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=ci, value=v)
            cell.font = font
            cell.border = bdr
            if ci >= 3:
                cell.alignment = Alignment(horizontal="center")
        row_num += 1

    # Footer
    row_num += 1
    ws.cell(row=row_num, column=1, value=f"Project: {proj_name}").font = Font(
        name="Arial", size=8, italic=True
    )
    ws.cell(row=row_num, column=6, value=f"Generated: {datetime.now().strftime('%d-%b-%Y')}").font = Font(
        name="Arial", size=8, italic=True
    )

    # Print settings
    ws.print_title_rows = "3:3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{proj_id}_cpm_schedule.xlsx"'},
    )


@app.get("/api/projects/{proj_id}/export/schedule.xer")
def export_xer(proj_id: str) -> Response:
    """Export current schedule as a P6-compatible .xer file."""
    import json as _json
    from xer_writer import build_xer_export

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT name FROM projects WHERE proj_id = ?", (proj_id,))
        proj_row = cur.fetchone()
        if not proj_row:
            raise HTTPException(404, "Project not found")

        cur.execute(
            "SELECT table_name, field_names, row_data FROM xer_raw_tables WHERE proj_id = ?",
            (proj_id,),
        )
        raw_tables = {}
        for row in cur.fetchall():
            tbl_name = row[0]
            fields = _json.loads(row[1])
            rows = _json.loads(row[2])
            raw_tables[tbl_name] = (fields, rows)

        if not raw_tables:
            raise HTTPException(400, "No original XER data stored. Re-import the .xer file to enable export.")

        acts = fetch_activities(conn, proj_id)
        rels = fetch_relationships(conn, proj_id)
    finally:
        conn.close()

    xer_content = build_xer_export(proj_id, raw_tables, acts, rels)

    return Response(
        content=xer_content.encode("utf-8"),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{proj_id}_export.xer"'},
    )


@app.get("/api/projects/{proj_id}/export/diagnostics.csv")
def export_diagnostics_csv(
    proj_id: str,
    hours_per_day: float = Query(8.0, ge=0.1, le=24.0),
) -> Response:
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        acts = fetch_activities(conn, proj_id)
        rels = fetch_relationships(conn, proj_id)
    finally:
        conn.close()

    rep = run_diagnostics(proj_id, acts, rels, hours_per_day=hours_per_day)

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["check", "severity", "task_id", "message"])
    w.writerow(
        [
            "summary",
            "info",
            "",
            f"total={rep.summary.total_activities}, critical_pct={rep.summary.critical_pct}, "
            f"open_starts={rep.summary.open_starts}, open_ends={rep.summary.open_ends}",
        ]
    )
    for f in rep.findings:
        w.writerow([f.check, f.severity, f.task_id or "", f.message])
    return Response(
        content=out.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{proj_id}_diagnostics.csv"'},
    )


@app.post("/api/projects/{proj_id}/baselines", response_model=BaselineInfo)
def save_baseline(proj_id: str, name: Optional[str] = None) -> BaselineInfo:
    """Save current schedule as a new baseline."""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        cur.execute(
            "SELECT COALESCE(MAX(baseline_number), 0) + 1 FROM baselines WHERE proj_id = ?",
            (proj_id,),
        )
        next_num = cur.fetchone()[0]
        cur.execute(
            "INSERT INTO baselines (proj_id, baseline_number, name) VALUES (?, ?, ?)",
            (proj_id, next_num, name),
        )
        baseline_id = cur.lastrowid
        cur.execute(
            """INSERT INTO baseline_activities
               (baseline_id, task_id, duration_hrs, early_start, early_finish,
                late_start, late_finish, total_float_hrs, is_critical)
               SELECT ?, task_id, duration_hrs, early_start, early_finish,
                      late_start, late_finish, total_float_hrs, is_critical
               FROM activities WHERE proj_id = ?""",
            (baseline_id, proj_id),
        )
        conn.commit()
        cur.execute(
            "SELECT id, proj_id, baseline_number, name, created_at FROM baselines WHERE id = ?",
            (baseline_id,),
        )
        row = cur.fetchone()
        return BaselineInfo(
            id=row[0],
            proj_id=row[1],
            baseline_number=row[2],
            name=row[3],
            created_at=row[4],
        )
    finally:
        conn.close()


@app.get("/api/projects/{proj_id}/baselines", response_model=List[BaselineInfo])
def list_baselines(proj_id: str) -> List[BaselineInfo]:
    """List all baselines for a project."""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        cur.execute(
            """SELECT id, proj_id, baseline_number, name, created_at
               FROM baselines WHERE proj_id = ? ORDER BY baseline_number""",
            (proj_id,),
        )
        rows = cur.fetchall()
        return [
            BaselineInfo(
                id=r[0], proj_id=r[1], baseline_number=r[2], name=r[3], created_at=r[4]
            )
            for r in rows
        ]
    finally:
        conn.close()


@app.get(
    "/api/projects/{proj_id}/baselines/{baseline_number}/compare",
    response_model=List[BaselineComparison],
)
def compare_baseline(proj_id: str, baseline_number: int) -> List[BaselineComparison]:
    """Compare current schedule with a baseline."""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        cur.execute(
            "SELECT id FROM baselines WHERE proj_id = ? AND baseline_number = ?",
            (proj_id, baseline_number),
        )
        bl_row = cur.fetchone()
        if not bl_row:
            raise HTTPException(404, "Baseline not found")
        baseline_id = bl_row[0]
        cur.execute(
            """SELECT ba.task_id, a.name,
                      ba.early_start, ba.early_finish, ba.duration_hrs,
                      a.early_start, a.early_finish, a.duration_hrs
               FROM baseline_activities ba
               LEFT JOIN activities a ON a.proj_id = ? AND a.task_id = ba.task_id
               WHERE ba.baseline_id = ?
               ORDER BY ba.task_id""",
            (proj_id, baseline_id),
        )
        rows = cur.fetchall()
        comparisons: List[BaselineComparison] = []
        for r in rows:
            bl_es, bl_ef, bl_dur = r[2], r[3], r[4]
            cur_es, cur_ef, cur_dur = r[5], r[6], r[7]
            start_var = (cur_es - bl_es) if (cur_es is not None and bl_es is not None) else None
            finish_var = (cur_ef - bl_ef) if (cur_ef is not None and bl_ef is not None) else None
            dur_var = (cur_dur - bl_dur) if (cur_dur is not None and bl_dur is not None) else None
            comparisons.append(
                BaselineComparison(
                    task_id=str(r[0]),
                    name=r[1] or "",
                    bl_early_start=bl_es,
                    bl_early_finish=bl_ef,
                    bl_duration_hrs=bl_dur,
                    cur_early_start=cur_es,
                    cur_early_finish=cur_ef,
                    cur_duration_hrs=cur_dur,
                    start_variance=start_var,
                    finish_variance=finish_var,
                    duration_variance=dur_var,
                )
            )
        return comparisons
    finally:
        conn.close()


@app.delete("/api/projects/{proj_id}/baselines/{baseline_number}")
def delete_baseline(proj_id: str, baseline_number: int) -> dict:
    """Delete a baseline."""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        cur.execute(
            "DELETE FROM baselines WHERE proj_id = ? AND baseline_number = ?",
            (proj_id, baseline_number),
        )
        if cur.rowcount == 0:
            raise HTTPException(404, "Baseline not found")
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@app.get("/api/projects/{proj_id}/calendar-dates")
def get_calendar_dates(proj_id: str) -> dict:
    """Return calendar-aware date strings for all activities."""
    from constants import REF_DATETIME
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")

        acts = fetch_activities(conn, proj_id)
        calendars = load_calendars(conn, proj_id)
        default_cal = get_default_calendar()
        ref_date = REF_DATETIME.date()

        date_map = {}
        for a in acts:
            tid = str(a["task_id"])
            cal_id = str(a.get("calendar_id") or "")
            cal = calendars.get(cal_id) or calendars.get(f"{proj_id}:{cal_id}") or default_cal

            entry = {}
            for fld in ("early_start", "early_finish", "late_start", "late_finish"):
                hrs = a.get(fld)
                if hrs is not None:
                    cal_date = hours_to_calendar_date(float(hrs), cal, ref_date)
                    entry[fld] = cal_date.isoformat()
                else:
                    entry[fld] = None
            date_map[tid] = entry

        return {
            "dates": date_map,
            "calendar_count": len(calendars),
            "using_default": len(calendars) == 0,
        }
    finally:
        conn.close()


@app.get("/api/projects/{proj_id}/resources")
def get_resources(proj_id: str) -> dict:
    """Resource summary: all resources and per-task assignments."""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")

        cur.execute(
            "SELECT rsrc_id, rsrc_short_name, rsrc_name, rsrc_type FROM resources WHERE proj_id = ? ORDER BY rsrc_name",
            (proj_id,),
        )
        resources = [{"rsrc_id": r[0], "short_name": r[1], "name": r[2], "type": r[3]} for r in cur.fetchall()]

        cur.execute(
            """SELECT tr.task_id, tr.rsrc_id, COALESCE(r.rsrc_name, tr.rsrc_name, tr.rsrc_id) as name,
                      tr.target_qty, tr.target_cost, tr.actual_cost, tr.remain_cost
               FROM task_resources tr
               LEFT JOIN resources r ON tr.rsrc_id = r.rsrc_id AND tr.proj_id = r.proj_id
               WHERE tr.proj_id = ?
               ORDER BY tr.task_id, name""",
            (proj_id,),
        )
        cols = [d[0] for d in cur.description]
        assignments = [dict(zip(cols, row)) for row in cur.fetchall()]

        return {"resources": resources, "assignments": assignments, "total_resources": len(resources), "total_assignments": len(assignments)}
    finally:
        conn.close()


@app.get("/api/projects/{proj_id}/activity-codes")
def get_activity_codes(proj_id: str) -> dict:
    """Return activity code types and their values for filtering/grouping."""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM projects WHERE proj_id = ?", (proj_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")

        cur.execute(
            "SELECT actv_code_type_id, actv_code_type, seq_num FROM activity_code_types WHERE proj_id = ? ORDER BY seq_num",
            (proj_id,),
        )
        types = [{"id": r[0], "name": r[1], "seq": r[2]} for r in cur.fetchall()]

        cur.execute(
            """SELECT ac.actv_code_type_id, ac.actv_code_id, ac.short_name, ac.actv_code_name, ac.color
               FROM activity_codes ac WHERE ac.proj_id = ?
               ORDER BY ac.actv_code_type_id, ac.seq_num""",
            (proj_id,),
        )
        values = {}
        for r in cur.fetchall():
            type_id = r[0]
            if type_id not in values:
                values[type_id] = []
            values[type_id].append({"id": r[1], "short_name": r[2], "name": r[3], "color": r[4]})

        return {"types": types, "values": values}
    finally:
        conn.close()


@app.get("/api/projects/{proj_id}/activities/{task_id}/codes")
def get_task_codes(proj_id: str, task_id: str) -> List[dict]:
    """Return activity codes assigned to a specific task."""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            """SELECT tac.actv_code_type_id,
                      COALESCE(act.actv_code_type, tac.actv_code_type_id) as type_name,
                      tac.actv_code_id,
                      COALESCE(ac.short_name, '') as short_name,
                      COALESCE(ac.actv_code_name, '') as code_name,
                      ac.color
               FROM task_activity_codes tac
               LEFT JOIN activity_code_types act ON tac.actv_code_type_id = act.actv_code_type_id
               LEFT JOIN activity_codes ac ON tac.actv_code_id = ac.actv_code_id
               WHERE tac.proj_id = ? AND tac.task_id = ?
               ORDER BY act.seq_num, ac.seq_num""",
            (proj_id, task_id),
        )
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]
    finally:
        conn.close()


@app.get("/api/health")
def health() -> dict:
    return {"status": "ok"}
